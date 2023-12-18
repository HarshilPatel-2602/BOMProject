from flask import Flask, render_template, request, flash , send_file
import openpyxl
import os
import re
import pandas as pd

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

app = Flask(__name__)
app.secret_key = 'super secret key'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

processed_file = '' 

def append_all_rows(result_array, final_workbook_path):
    # Create the final workbook
    final_workbook = openpyxl.load_workbook(final_workbook_path)
    final_sheet = final_workbook['ModBOM']
    # Iterate over all sheet types
    for row_array in result_array:
        final_sheet.append(list(row_array)) 
    final_workbook.save(final_workbook_path)
    final_workbook.close()

def extract_substring(input_string):
    # Use regular expression to extract the desired substring
    match = re.search(r'[A-Z].*\d', input_string)
    if match:
        return match.group()
    else:
        return None

def sort_and_replace_final_workbook(final_workbook_path, sheet_type_priorities):
    # Load the final workbook
    final_workbook = openpyxl.load_workbook(final_workbook_path)
    final_sheet = final_workbook['ModBOM']
    # Extract the header row
    header_row = final_sheet[1]
    header_values = [cell.value for cell in header_row]
    # Sort the rows based on the sheet type priorities
    sorted_rows = sorted(final_sheet.iter_rows(min_row=2, values_only=True), key=lambda row: sheet_type_priorities.get(row[0], float('inf')))
    # Clear existing data in the final sheet
    final_sheet.delete_rows(2, final_sheet.max_row)
    # Append the sorted rows to the final sheet
    for row_data in sorted_rows:
        final_sheet.append(row_data)
    # Save the changes
    final_workbook.save(final_workbook_path)
    final_workbook.close()

def process_file(file1,file2,type):
    if 'file1' not in request.files or 'file2' not in request.files :
        flash('No file part')
        return "error"
    file1 = request.files['file1']
    file2 = request.files['file2']
    # If the user does not select a file, the browser submits an empty file without a filename.
    if file1.filename == '' or file2.filename == '' :
        flash('No selected file')
        return "error: no selected file"
    # Save the files to the upload folder
    file1_path = os.path.join(app.config['UPLOAD_FOLDER'], file1.filename)
    file2_path = os.path.join(app.config['UPLOAD_FOLDER'], file2.filename)
    file1.save(file1_path)
    file2.save(file2_path)
    # Now you can use file_path to load the workbook using openpyxl
    workbook = openpyxl.load_workbook(file1_path)
    # Iterate over all sheets
    result_array = []
    for sheet_name in workbook.sheetnames :
        sheet_input = workbook[sheet_name] ;
        # Add further code here for processing the selected workbook
        all_rows_data = []
        for row in sheet_input.iter_rows(min_row=2, max_row=sheet_input.max_row, values_only=True):
            all_rows_data.append(row) ;
        #print("All Rows Data:", all_rows_data[0])
        for row_data in all_rows_data:
            BaseMode = sheet_name ;
            Module = BaseMode + '-00' ;
            Item = row_data[1] ;
            BaseItem = extract_substring(Item) ;
            # Create an array for the current row
            current_row_array = [type,row_data[0], Module, BaseMode, Item, BaseItem] + list(row_data[2:])
            # Append the array for the current row to the result_array
            result_array.append(current_row_array)
    # Close the workbook when done
    workbook.close()
    # final = 'static/EPA24 JES Mixer Shortege Sheet_TLA UPDATED & WITHOUT DUPLICATE_MIXER BOM(11_6_23).xlsx' 
    append_all_rows(result_array,file2_path) ;
    # clear_all_row(sheet_type_to_workbook) ;
    sheet_type_priorities = {'DPF': 1, 'DOC': 2, 'SCR': 3, 'MIXER': 4}
    sort_and_replace_final_workbook(file2_path, sheet_type_priorities)
    return file2_path

@app.route("/") 
def home():
    return render_template("index.html")

@app.route("/edit", methods=["GET", "POST"])
def edit():
    if request.method == "POST":
        type = request.form.get("type")
        # Specify the path to your file
        processed_file = process_file(request.files['file1'],request.files['file2'],type)
        filename = 'final.xlsx' 
        return send_file(processed_file, as_attachment=True, download_name=filename)
    else:
        return render_template("index.html")


if __name__ == '__main__':
    app.run(debug=True,host='0.0.0.0')